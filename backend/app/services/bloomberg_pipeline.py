"""
Bloomberg Excel Export Pipeline — Values Sheet Parser.

Parses JMWFM_Bloomberg_YYYY-MM-DD.xlsx workbooks exported from Bloomberg Terminal.
Reads the Values sheet only (Fundamentals sheet shows #NAME? outside terminal — expected).

Recommended cadence: Daily upload after market close.
"""

import logging
import re
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl

from app.services.supabase_client import get_supabase

logger = logging.getLogger("wasden_watch.bloomberg_pipeline")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Column mapping: Excel header -> DB column
COLUMN_MAP = {
    "Current Price": "price",
    "Market Cap": "market_cap",
    "P/E Ratio (Trailing)": "trailing_pe",
    "P/E Ratio (Forward)": "forward_pe",
    "EPS (Current)": "eps",
    "PEG Ratio": "peg_ratio",
    "Free Cash Flow": "fcf",
    "FCF Yield": "fcf_yield",
    "EBITDA Margin": "ebitda_margin",
    "ROE": "roe",
    "ROC": "roc",
    "Gross Margin": "gross_margin",
    "Operating Margin": "operating_margin",
    "Net Margin": "net_margin",
    "Current Ratio": "current_ratio",
    "Quick Ratio": "quick_ratio",
    "Debt to Equity": "debt_to_equity",
    "Revenue Growth YoY": "revenue_growth",
    "EBITDA/Interest": "ebitda_interest_coverage",
    "Cash Conversion Cycle": "ccc",
    "Short Interest": "short_interest",
}

# Known ADR tickers
ADR_TICKERS = {"TSM"}

# Bloomberg error patterns -> typed error codes
ERROR_PATTERNS = {
    r"#N/A Invalid Field": "N/A_INVALID_FIELD",
    r"#N/A Field Not Applicable": "N/A_NOT_APPLICABLE",
    r"#N/A N/A": "N/A_NA",
    r"#N/A": "N/A_GENERIC",
    r"#VALUE!": "VALUE_ERROR",
    r"#NAME\?": "NAME_ERROR",
}

# Freshness thresholds
FRESH_HOURS = 24
RECENT_DAYS = 7
STALE_DAYS = 30


# ---------------------------------------------------------------------------
# Freshness grading
# ---------------------------------------------------------------------------

def compute_freshness(pull_date: date, reference_date: Optional[date] = None) -> str:
    """Compute data freshness grade based on age.

    FRESH: <24 hours -- full weight
    RECENT: 1-7 days -- full weight, flagged in logs
    STALE: 7-30 days -- weight reduced by 50%
    EXPIRED: >30 days -- excluded from live decisions

    Delegates to shared freshness module and returns .value string
    for backward compatibility with existing callers.
    """
    from app.services.freshness import compute_freshness as _compute
    return _compute(pull_date, reference_date).value


# ---------------------------------------------------------------------------
# Error detection
# ---------------------------------------------------------------------------

def classify_error(raw_value: str) -> Optional[str]:
    """Classify a Bloomberg cell error. Returns typed error code or None."""
    if not isinstance(raw_value, str):
        return None
    val = raw_value.strip()
    # Check more specific patterns first
    for pattern, error_type in ERROR_PATTERNS.items():
        if re.search(pattern, val):
            return error_type
    return None


def parse_numeric(raw_value) -> tuple[Optional[float], Optional[str]]:
    """Parse a cell value to float. Returns (value, error_type).

    If the cell is a Bloomberg error string, returns (None, error_type).
    If the cell is empty/None, returns (None, None).
    If numeric, returns (float_value, None).
    """
    if raw_value is None:
        return None, None

    if isinstance(raw_value, (int, float)):
        return float(raw_value), None

    if isinstance(raw_value, str):
        error_type = classify_error(raw_value)
        if error_type:
            return None, error_type
        # Try parsing as number
        cleaned = raw_value.strip().replace(",", "")
        if not cleaned:
            return None, None
        try:
            return float(cleaned), None
        except ValueError:
            logger.warning(f"Unparseable value: {raw_value!r}")
            return None, "PARSE_ERROR"

    return None, None


def extract_ticker(bloomberg_ticker: str) -> str:
    """Extract plain ticker from Bloomberg format.

    'NVDA US Equity' -> 'NVDA'
    'NVDA' -> 'NVDA'
    """
    if not bloomberg_ticker:
        return ""
    return bloomberg_ticker.strip().split()[0].upper()


# ---------------------------------------------------------------------------
# Excel parser
# ---------------------------------------------------------------------------

def parse_bloomberg_excel(file_path: str | Path) -> dict:
    """Parse a Bloomberg Excel export and return structured data.

    Args:
        file_path: Path to JMWFM_Bloomberg_YYYY-MM-DD.xlsx

    Returns:
        dict with keys:
            - pull_date: date extracted from filename or sheet
            - ticker_data: list of dicts, one per ticker row
            - validation: BloombergValidationReport-compatible dict
    """
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Bloomberg file not found: {file_path}")

    # Extract pull date from filename
    pull_date = _extract_date_from_filename(file_path.name)

    # Load workbook -- Values sheet only
    wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)

    sheet_name = None
    for name in wb.sheetnames:
        if "values" in name.lower():
            sheet_name = name
            break

    if sheet_name is None:
        # Fall back to first sheet
        sheet_name = wb.sheetnames[0]
        logger.warning(f"No 'Values' sheet found; using first sheet: {sheet_name}")

    ws = wb[sheet_name]

    # Read headers from first row
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ValueError("Empty worksheet -- no data found")

    headers = [str(h).strip() if h else "" for h in rows[0]]

    # Build column index mapping
    col_indices = {}
    ticker_col = None
    date_col = None

    for i, header in enumerate(headers):
        if header.lower() in ("ticker", "security"):
            ticker_col = i
        elif header.lower() == "date":
            date_col = i
        elif header in COLUMN_MAP:
            col_indices[header] = i

    if ticker_col is None:
        raise ValueError("No 'Ticker' column found in Values sheet")

    # Parse each data row
    ticker_data = []
    ticker_results = []

    for row_idx, row in enumerate(rows[1:], start=2):
        if not row or not row[ticker_col]:
            continue

        raw_ticker = str(row[ticker_col]).strip()
        if not raw_ticker:
            continue

        ticker = extract_ticker(raw_ticker)
        is_adr = ticker in ADR_TICKERS

        # Parse row date if available
        row_date = pull_date
        if date_col is not None and row[date_col]:
            try:
                if isinstance(row[date_col], datetime):
                    row_date = row[date_col].date()
                elif isinstance(row[date_col], date):
                    row_date = row[date_col]
            except Exception:
                pass

        # Parse each metric column
        record = {
            "ticker": ticker,
            "pull_date": str(row_date),
            "is_adr": is_adr,
        }
        errors = []
        fields_parsed = 0
        fields_errored = 0

        for header, db_col in COLUMN_MAP.items():
            idx = col_indices.get(header)
            if idx is None or idx >= len(row):
                continue

            value, error_type = parse_numeric(row[idx])
            record[db_col] = value
            fields_parsed += 1

            if error_type:
                fields_errored += 1
                errors.append({
                    "field_name": db_col,
                    "raw_value": str(row[idx]) if row[idx] is not None else "",
                    "error_type": error_type,
                })

        # Compute freshness
        freshness = compute_freshness(row_date)
        record["freshness_grade"] = freshness
        record["is_error"] = fields_errored > 0
        record["error_fields"] = errors if errors else None

        ticker_data.append(record)
        ticker_results.append({
            "ticker": ticker,
            "success": fields_errored == 0,
            "fields_parsed": fields_parsed,
            "fields_errored": fields_errored,
            "errors": errors,
            "freshness": freshness,
        })

    successful = sum(1 for r in ticker_results if r["success"])
    failed = len(ticker_results) - successful

    validation = {
        "filename": file_path.name,
        "pull_date": str(pull_date),
        "total_tickers": len(ticker_results),
        "successful_tickers": successful,
        "failed_tickers": failed,
        "ticker_results": ticker_results,
        "uploaded_to_supabase": False,
        "errors_summary": [
            f"{r['ticker']}: {r['fields_errored']} field errors"
            for r in ticker_results
            if not r["success"]
        ],
    }

    return {
        "pull_date": pull_date,
        "ticker_data": ticker_data,
        "validation": validation,
    }


def _extract_date_from_filename(filename: str) -> date:
    """Extract date from JMWFM_Bloomberg_YYYY-MM-DD.xlsx filename."""
    match = re.search(r"(\d{4}-\d{2}-\d{2})", filename)
    if match:
        return date.fromisoformat(match.group(1))
    # Fallback to today
    logger.warning(f"Could not extract date from filename: {filename}. Using today.")
    return date.today()


# ---------------------------------------------------------------------------
# Supabase upload
# ---------------------------------------------------------------------------

def upload_to_supabase(ticker_data: list[dict]) -> int:
    """Upsert parsed Bloomberg data to Supabase bloomberg_fundamentals table.

    Returns the number of rows upserted.
    """
    if not ticker_data:
        return 0

    client = get_supabase()
    upserted = 0

    for record in ticker_data:
        # Build the row for Supabase -- exclude None values for cleaner upsert
        row = {k: v for k, v in record.items() if v is not None}

        # Serialize error_fields as JSON
        if "error_fields" in row and row["error_fields"]:
            # Already a list of dicts, Supabase handles JSONB
            pass
        else:
            row.pop("error_fields", None)

        try:
            client.table("bloomberg_fundamentals").upsert(
                row,
                on_conflict="ticker,pull_date",
            ).execute()
            upserted += 1
        except Exception as e:
            logger.error(f"Failed to upsert {record.get('ticker')}: {e}")

    return upserted


# ---------------------------------------------------------------------------
# Main pipeline entry point
# ---------------------------------------------------------------------------

def run_bloomberg_pipeline(
    file_path: str | Path,
    upload: bool = True,
) -> dict:
    """Run the full Bloomberg export pipeline.

    1. Parse Excel workbook (Values sheet)
    2. Validate all fields, classify errors
    3. Compute freshness grades
    4. Upload to Supabase (if upload=True)
    5. Return validation report

    Args:
        file_path: Path to JMWFM_Bloomberg_YYYY-MM-DD.xlsx
        upload: Whether to upload to Supabase (default True)

    Returns:
        Validation report dict (BloombergValidationReport-compatible)
    """
    logger.info(f"Starting Bloomberg pipeline for: {file_path}")

    # Step 1-3: Parse, validate, compute freshness
    result = parse_bloomberg_excel(file_path)
    validation = result["validation"]
    ticker_data = result["ticker_data"]

    logger.info(
        f"Parsed {validation['total_tickers']} tickers: "
        f"{validation['successful_tickers']} clean, "
        f"{validation['failed_tickers']} with errors"
    )

    # Step 4: Upload to Supabase
    if upload and ticker_data:
        try:
            upserted = upload_to_supabase(ticker_data)
            validation["uploaded_to_supabase"] = True
            logger.info(f"Upserted {upserted} rows to Supabase")
        except Exception as e:
            logger.error(f"Supabase upload failed: {e}")
            validation["errors_summary"].append(f"Supabase upload failed: {e}")
    elif not upload:
        logger.info("Skipping Supabase upload (upload=False)")

    return validation


def get_freshness_report() -> dict:
    """Query Supabase for the latest Bloomberg data freshness per ticker.

    Returns a FreshnessReport-compatible dict.
    """
    client = get_supabase()
    today = date.today()

    # Get the most recent pull for each ticker
    response = (
        client.table("bloomberg_fundamentals")
        .select("ticker, pull_date, freshness_grade")
        .order("pull_date", desc=True)
        .execute()
    )

    # Deduplicate -- keep most recent per ticker
    seen = {}
    for row in response.data:
        ticker = row["ticker"]
        if ticker not in seen:
            seen[ticker] = row

    tickers = []
    counts = {"FRESH": 0, "RECENT": 0, "STALE": 0, "EXPIRED": 0}

    for ticker, row in sorted(seen.items()):
        pull_date = date.fromisoformat(row["pull_date"])
        freshness = compute_freshness(pull_date, today)
        days_old = (today - pull_date).days

        tickers.append({
            "ticker": ticker,
            "pull_date": str(pull_date),
            "freshness": freshness,
            "days_old": days_old,
        })
        counts[freshness] = counts.get(freshness, 0) + 1

    return {
        "report_date": str(today),
        "tickers": tickers,
        "fresh_count": counts["FRESH"],
        "recent_count": counts["RECENT"],
        "stale_count": counts["STALE"],
        "expired_count": counts["EXPIRED"],
    }
